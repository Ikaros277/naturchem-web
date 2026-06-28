"""Generate XLSX checklist for image delivery."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "fotografie-dodavka-checklist.xlsx"

SECTOR_ROWS = [
    ("A", "lakovny", "Lakovny a povrchové úpravy", "Lakovací box, stříkání, odsávání", "poradna/paint-spray.jpg"),
    ("A", "galvanovny", "Galvanovny a tryskání", "Galvanická linka, kádě, odsávání", "mereni-emisi.webp (SDÍLÍ)"),
    ("A", "svarovny", "Svařovny", "Svařovna, dým, odsávání", "forge-worker.jpg"),
    ("A", "tiskarny-textilie", "Tiskárny a textilie", "Tisk, kašírování, lepidla", "poradna/automotive.jpg (SDÍLÍ)"),
    ("A", "drevozpracujici", "Dřevozpracující provozy", "Truhlárna, piliny, CNC", "poradna/woodworking.jpg"),
    ("A", "sklarstvi", "Sklářství", "Sklářská pec, výduch", "industrial-plant.jpg (SDÍLÍ)"),
    ("A", "automotive", "Automotive", "Montážní linka", "poradna/automotive.jpg (SDÍLÍ)"),
    ("A", "kotelny", "Kotelny", "Plynová/biomasová kotelna", "poradna/boiler-room.jpg"),
    ("A", "bioplyn-biometan", "Bioplyn a biometan", "Bioplynka, kogenerace", "ghg-overovani.webp (SDÍLÍ)"),
    ("A", "teplarny", "Teplárny", "Heizwerk, více kotlů", "industrial-plant.jpg (SDÍLÍ)"),
    ("A", "krematoria", "Krematoria", "Krematorium / pec", "mereni-emisi.webp (SDÍLÍ)"),
    ("A", "skladky-odpady", "Skládky", "Skládka, manipulace", "poradna/waste-landfill.jpg"),
    ("A", "odpady-recyklace", "Recyklace", "Drtič, třídička", "poradna/recycling.jpg"),
    ("A", "kompostarny", "Kompostárny", "Kompostování, bioodpady", "poradna/recycling.jpg (SDÍLÍ)"),
    ("A", "stavebni-zamery", "Stavební záměry", "Staveniště, terén", "homepage-eia.webp"),
    ("A", "kamenolomy", "Kamenolomy", "Lom, drcení, prach", "industrial-plant.jpg (SDÍLÍ)"),
    ("A", "zemedelske-provozy", "Zemědělské provozy", "Stáj, farma", "mereni-emisi.webp (SDÍLÍ)"),
    ("A", "susarny-zemedelstvi", "Sušárny zemědělství", "Sušárna obilí/biomasy", "mereni-emisi.webp (SDÍLÍ)"),
    ("A", "potravinarstvi", "Potravinářství", "Výrobní linka, chlazení", "pracovni-prostredi.webp (SDÍLÍ)"),
    ("A", "cov-kaly", "ČOV a kaly", "ČOV, kalové hospodářství", "odborne-posudky.webp (SDÍLÍ)"),
    ("A", "susarny-kalu", "Sušárny kalů", "Sušárna kalů", "odborne-posudky.webp (SDÍLÍ)"),
    ("A", "pyrolyza-kalu", "Pyrolýza kalů", "Pyrolýza, čištění spalin", "odborne-posudky.webp (SDÍLÍ)"),
    ("A", "cisteni-spalin", "Čištění spalin", "Filtr, měření před/za", "mereni-emisi.webp (SDÍLÍ)"),
    ("A", "tepelna-cerpadla-vzt", "TČ a VZT", "Venkovní jednotka TČ/VZT", "poradna/hvac-units.jpg"),
    ("A", "verejne-budovy", "Veřejné budovy", "Škola, nemocnice, úřad", "poradna/hvac-units.jpg (SDÍLÍ)"),
    ("A", "laboratore", "Laboratoře", "Laboratoř, čistý prostor", "pracovni-prostredi.webp (SDÍLÍ)"),
    ("A", "lesnictvi-doprava", "Lesnictví a doprava", "Stroj v terénu", "mereni-hluku.webp"),
    ("A", "ispop-evidence", "ISPOP", "Evidence, formuláře", "ispop.webp"),
    ("A", "ghg-cnr", "GHG/CNR", "Reporting skleníkových plynů", "ghg-overovani.webp (SDÍLÍ)"),
    ("A", "provozni-rady", "Provozní řády", "Dokumentace u zdroje", "provozni-rady.webp"),
    ("A", "odborne-posudky-povoleni", "Posudky a povolení", "Posudek, schémata", "odborne-posudky.webp (SDÍLÍ)"),
]

CASE_STUDY_BRIEFS: dict[str, str] = {
    "mereni-emisi-plynova-kotelna-verejna-budova": "Kotelna ve veřejné budově",
    "mereni-emisi-kotel-biomasa": "Kotel na dřevo/štěpku",
    "mereni-emisi-novy-kotel": "Výměna kotle, nový hořák",
    "mereni-emisi-plynovy-horak": "Plynový hořák u technologie",
    "mereni-emisi-voc-tisk": "Tiskárna, VOC, lepidla",
    "mereni-emisi-filtr-aktivni-uhli": "Filtr s aktivním uhlím",
    "mereni-emisi-tryskani": "Pískování/tryskání",
    "pracovni-prostredi-chemicke-latky": "Chemické látky na lince",
    "pracovni-prostredi-mikroklima": "Tepelná zátěž u pece/lisu",
    "pracovni-prostredi-hluk-hala": "Hluk ve výrobní hale",
    "pracovni-prostredi-osvetleni": "Měření osvětlení",
    "pracovni-prostredi-komplex": "Kombinace faktorů PP",
    "mereni-vibraci-ruce": "Vibrace rukou (HAV)",
    "mereni-vibraci-cele-telo": "Vibrace celého těla (WBV)",
    "hlukova-studie-vyrobni-areal": "Hlukový areál shora",
    "mereni-hluku-dieselgenerator": "Dieselagregát",
    "posudek-tryskaci-zarizeni": "Tryskací zařízení",
    "aktualizace-provozniho-radu": "Aktualizace provozního řádu",
    "vyjadreni-kontrola-cizp": "Kontrola ČIŽP",
    "oprava-ispop": "Oprava ISPOP hlášení",
    "pfas-povrchove-upravy": "PFAS, povrchové úpravy",
    "bezpecnostni-list-chemicky-vyrobek": "Bezpečnostní list výrobku",
    "mereni-emisi-pradelna": "Prádelna, výduch",
    "kontrolni-mereni-pozadavek-khs": "Kontrolní měření pro KHS",
}

SECTORS_MAP = {r[1]: r[4] for r in SECTOR_ROWS}
CAT_MAP = {
    "mereni-emisi": "mereni-emisi.webp",
    "pracovni-prostredi": "pracovni-prostredi.webp",
    "hluk-vibrace": "mereni-hluku.webp",
    "rozptylove-studie": "rozptylove-studie.webp",
    "hlukove-studie": "hlukove-studie.webp",
    "eia-jes": "eia.webp",
    "odborne-posudky": "odborne-posudky.webp",
    "provozni-rady": "provozni-rady.webp",
    "ispop-evidence": "ispop.webp",
    "specificke-technologie": "provozy-a-technologie.webp",
}


def parse_case_studies() -> list[dict[str, str]]:
    text = (ROOT / "src/data/case-studies-cs.ts").read_text(encoding="utf-8")
    blocks = re.split(r"\n  \{\n", text)
    studies: list[dict[str, str]] = []
    for block in blocks[1:]:
        id_m = re.search(r'id: "([^"]+)"', block)
        if not id_m:
            continue
        title_m = re.search(r'title: "([^"]+)"', block)
        cat_m = re.search(r'categoryId: "([^"]+)"', block)
        sec_m = re.search(r'sectorId: "([^"]+)"', block)
        studies.append(
            {
                "id": id_m.group(1),
                "title": title_m.group(1) if title_m else "",
                "categoryId": cat_m.group(1) if cat_m else "",
                "sectorId": sec_m.group(1) if sec_m else "",
            }
        )
    return studies


def current_case_image(study: dict[str, str]) -> str:
    if study["sectorId"]:
        return SECTORS_MAP.get(study["sectorId"], "provozy-a-technologie.webp")
    return CAT_MAP.get(study["categoryId"], "typicke-zakazky.webp")


def setup_sheet(ws, headers: list[str], widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="1A5C3A")
    header_font = Font(color="FFFFFF", bold=True)
    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def write_rows(ws, rows: list[tuple], start_row: int = 2) -> None:
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet 1: Instructions
    ws0 = wb.active
    ws0.title = "Navod"
    instructions = [
        "CHECKLIST DODÁVKY FOTOGRAFIÍ — NATURCHEM web",
        "",
        "Formát: WebP nebo JPG, min. 1200 px šířka, poměr 16:9 (dlaždice) nebo 3:2 (hero).",
        "",
        "List Provozy — 31 fotek do public/hero/provozy/{id}.webp",
        "List Zakazky-bez-sectorId — priorita B (22 ks)",
        "List Zakazky-vse — volitelně 50 unikátních fotek (priorita C)",
        "List Poradna — rozšíření fotek u článků (priorita D)",
        "",
        "Sloupec Dodano: vyplňte ANO až po dodání souboru.",
        "Po dodání souborů napojím mapování v kódu (sector-group-visuals, case-study-visuals).",
    ]
    for i, line in enumerate(instructions, start=1):
        ws0.cell(row=i, column=1, value=line)
    ws0.column_dimensions["A"].width = 100

    # Sheet 2: Sectors
    ws1 = wb.create_sheet("Provozy")
    h1 = [
        "Priorita",
        "ID provozu",
        "Název provozu",
        "Co má být na fotce",
        "Cílová cesta souboru",
        "Aktuálně používaná fotka",
        "Dodáno (ANO/NE)",
        "Poznámka",
    ]
    setup_sheet(ws1, h1, [10, 28, 32, 42, 40, 36, 14, 24])
    rows1 = [
        (
            prio,
            sid,
            name,
            brief,
            f"public/hero/provozy/{sid}.webp",
            current,
            "NE",
            "",
        )
        for prio, sid, name, brief, current in SECTOR_ROWS
    ]
    write_rows(ws1, rows1)

    studies = parse_case_studies()
    no_sector = [s for s in studies if not s["sectorId"]]

    # Sheet 3: Case studies without sectorId
    ws2 = wb.create_sheet("Zakazky-bez-sectorId")
    h2 = [
        "Priorita",
        "ID zakázky",
        "Název zakázky",
        "Kategorie",
        "Co má být na fotce",
        "Cílová cesta souboru",
        "Aktuálně používaná fotka",
        "Dodáno (ANO/NE)",
        "Poznámka",
    ]
    setup_sheet(ws2, h2, [10, 34, 44, 22, 36, 42, 34, 14, 24])
    rows2 = [
        (
            "B",
            s["id"],
            s["title"],
            s["categoryId"],
            CASE_STUDY_BRIEFS.get(s["id"], "Viz název zakázky"),
            f"public/hero/case-studies/{s['id']}.webp",
            current_case_image(s),
            "NE",
            "",
        )
        for s in no_sector
    ]
    write_rows(ws2, rows2)

    # Sheet 4: All case studies
    ws3 = wb.create_sheet("Zakazky-vse")
    h3 = [
        "Priorita",
        "ID zakázky",
        "Název zakázky",
        "sectorId",
        "Kategorie",
        "Co má být na fotce",
        "Cílová cesta souboru",
        "Aktuálně používaná fotka",
        "Dodáno (ANO/NE)",
    ]
    setup_sheet(ws3, h3, [10, 34, 44, 24, 22, 36, 42, 34, 14])
    rows3 = [
        (
            "C" if s["sectorId"] else "B",
            s["id"],
            s["title"],
            s["sectorId"] or "—",
            s["categoryId"],
            CASE_STUDY_BRIEFS.get(s["id"], "Specifická scéna k dané zakázce"),
            f"public/hero/case-studies/{s['id']}.webp",
            current_case_image(s),
            "NE",
        )
        for s in studies
    ]
    write_rows(ws3, rows3)

    # Sheet 5: Poradna summary
    ws4 = wb.create_sheet("Poradna")
    h4 = [
        "Priorita",
        "Téma / slug článku",
        "Doporučený obsah fotky",
        "Cílová cesta",
        "Dodáno (ANO/NE)",
    ]
    setup_sheet(ws4, h4, [10, 50, 48, 44, 14])
    poradna = (ROOT / "src/lib/poradna-hero-images.ts").read_text(encoding="utf-8")
    article_slugs = re.findall(r'"([^"]+)":', poradna)
    theme_per_slug = re.findall(r'"([^"]+)":\s*"([^"]+)"', poradna)
    rows4 = [
        (
            "D",
            slug,
            f"Téma: {theme}",
            f"public/hero/poradna/clanky/{slug}.webp",
            "NE",
        )
        for slug, theme in theme_per_slug
        if slug != "poradnaArticleHeroThemes"
    ]
    write_rows(ws4, rows4)

    wb.save(OUT)
    print(f"Written {OUT}")
    print(f"Sectors: {len(SECTOR_ROWS)}, case studies: {len(studies)}, without sectorId: {len(no_sector)}")


if __name__ == "__main__":
    main()
